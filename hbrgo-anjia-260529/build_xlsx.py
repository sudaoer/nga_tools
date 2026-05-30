from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WORK_DIR = Path(__file__).resolve().parent
ANJIA_DIR = WORK_DIR / "anjia"
XLSX_PATH = WORK_DIR / "hbrgo-anjia-260529.xlsx"


def _read_json(path: Path) -> dict[str, Any]:
    raw_data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw_data, dict):
        raise ValueError(f"JSON顶层必须是对象：{path}")
    return raw_data


def _record_paths() -> list[Path]:
    if not ANJIA_DIR.exists():
        raise FileNotFoundError(
            f"未找到安价JSON目录：{ANJIA_DIR}。请先运行extract_anjia.py。"
        )
    return sorted(ANJIA_DIR.glob("anjia_*.json"))


def _source_value(record: dict[str, Any], key: str) -> object:
    source_post = record.get("source_post")
    if not isinstance(source_post, dict):
        return ""
    return source_post.get(key, "")


def _source_int(record: dict[str, Any], key: str) -> int:
    value = _source_value(record, key)
    if type(value) is int:
        return value
    return -1


def _source_str(record: dict[str, Any], key: str) -> str:
    value = _source_value(record, key)
    if isinstance(value, str):
        return value
    return ""


def _record_order_key(record: dict[str, Any]) -> tuple[str, int, int]:
    return (
        _source_str(record, "postdate"),
        _source_int(record, "lou"),
        _source_int(record, "pid"),
    )


def _last_records_by_uid(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records_by_uid: dict[int, dict[str, Any]] = {}
    for record in records:
        uid = _source_int(record, "uid")
        if uid < 0:
            continue
        previous = records_by_uid.get(uid)
        if previous is None or _record_order_key(record) > _record_order_key(previous):
            records_by_uid[uid] = record

    return sorted(records_by_uid.values(), key=_record_order_key)


def _xlsx_candidate(record: dict[str, Any]) -> bool:
    flags = record.get("flags")
    if not isinstance(flags, list):
        return True
    return not any(str(flag).startswith("forbidden_") for flag in flags)


def _flags(record: dict[str, Any]) -> str:
    flags = record.get("flags")
    if not isinstance(flags, list):
        return ""
    xlsx_flags = [str(flag) for flag in flags if flag != "duplicate_user_submission"]
    return ", ".join(xlsx_flags)


def _notes(record: dict[str, Any]) -> str:
    notes = record.get("notes")
    if not isinstance(notes, str):
        return ""
    duplicate_note = "同一用户的后续投稿；按规则首条为采纳候选。"
    return notes.replace(duplicate_note, "").strip()


def _row(record: dict[str, Any], number: Optional[int]) -> list[object]:
    return [
        number if number is not None else "",
        "是" if _xlsx_candidate(record) else "否",
        record.get("user_submission_index", ""),
        _source_value(record, "uid"),
        _source_value(record, "lou"),
        _source_value(record, "page"),
        _source_value(record, "pid"),
        _source_value(record, "postdate"),
        record.get("plain_text", ""),
        _flags(record),
        _notes(record),
    ]


def _rows(records: list[dict[str, Any]], column_count: int) -> list[list[object]]:
    valid_records = [record for record in records if _xlsx_candidate(record)]
    invalid_records = [record for record in records if not _xlsx_candidate(record)]

    rows: list[list[object]] = []
    for index, record in enumerate(valid_records, start=1):
        rows.append(_row(record, index))

    if invalid_records:
        rows.append([""] * column_count)
        for record in invalid_records:
            rows.append(_row(record, None))

    return rows


def _autosize_columns(sheet: object, widths: dict[int, int]) -> None:
    for column_index, width in widths.items():
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = width


def main() -> None:
    all_records = [_read_json(path) for path in _record_paths()]
    records = _last_records_by_uid(all_records)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "安价投稿"

    headers = [
        "编号",
        "采纳候选",
        "同用户第几条",
        "UID",
        "原楼层",
        "页码",
        "PID",
        "发帖时间",
        "全文",
        "标记",
        "备注",
    ]
    sheet.append(headers)
    for row in _rows(records, len(headers)):
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autosize_columns(
        sheet,
        {
            1: 8,
            2: 10,
            3: 14,
            4: 12,
            5: 10,
            6: 8,
            7: 14,
            8: 18,
            9: 80,
            10: 24,
            11: 36,
        },
    )

    workbook.save(XLSX_PATH)
    print(f"已生成xlsx：{XLSX_PATH}")
    print(f"逐条JSON记录数：{len(all_records)}")
    print(f"xlsx导出记录数：{len(records)}")
    print(f"xlsx有效内容数：{sum(1 for record in records if _xlsx_candidate(record))}")
    print(f"xlsx未采纳内容数：{sum(1 for record in records if not _xlsx_candidate(record))}")


if __name__ == "__main__":
    main()
