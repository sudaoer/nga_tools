from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
except ModuleNotFoundError as exc:
    raise SystemExit(
        "openpyxl is required. Run this script with `pixi run python ...` or install dependencies first."
    ) from exc


DEFAULT_INPUT = Path("anjia-viewer/public/data/anchors_43877379.json")
DEFAULT_MAX_COLUMN_WIDTH = 72
EXCEL_CELL_LIMIT = 32767
TRUNCATION_SUFFIX = "\n...[truncated for Excel cell limit]"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
SHEET_TITLE_LIMIT = 31
TOPIC_FIELD_ORDER: dict[str, list[str]] = {
    "theme1_travel": ["人物", "地点", "事件"],
    "theme2_training": ["老师", "主要属性培养", "事件"],
    "theme2_guest": ["客人", "事件"],
    "theme3_engagement": ["人物", "地点", "事件"],
    "theme4_carriage_name": ["名字"],
    "qa": ["问题"],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export anchor counter JSON to an XLSX workbook.")
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the generated anchors JSON file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Path to the output XLSX file. Defaults to the input path with an .xlsx suffix.",
    )
    parser.add_argument(
        "--max-column-width",
        type=int,
        default=DEFAULT_MAX_COLUMN_WIDTH,
        help="Maximum Excel column width for auto-sized columns.",
    )
    return parser.parse_args()


def load_payload(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def default_output_path(input_path: Path) -> Path:
    return input_path.with_suffix(".xlsx")


def truncate_text(value: str) -> str:
    if len(value) <= EXCEL_CELL_LIMIT:
        return value
    allowed = EXCEL_CELL_LIMIT - len(TRUNCATION_SUFFIX)
    return value[:allowed] + TRUNCATION_SUFFIX


def compact_json(value: Any) -> str:
    return truncate_text(json.dumps(value, ensure_ascii=False, separators=(",", ": ")))


def join_lines(items: list[Any]) -> str:
    values = [str(item) for item in items if item is not None and str(item) != ""]
    return truncate_text("\n".join(values))


def format_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return truncate_text(value)
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, list):
        if all(not isinstance(item, (dict, list)) for item in value):
            return join_lines(value)
        return compact_json(value)
    if isinstance(value, dict):
        return compact_json(value)
    return truncate_text(str(value))


def post_url(pid: Any) -> str:
    if pid is None or str(pid).strip() == "":
        return ""
    return f"https://ngabbs.com/read.php?pid={str(pid).strip()}&opt=128"


def source_lous_text(entry: dict[str, Any]) -> str:
    source_lous = entry.get("source_lous")
    if isinstance(source_lous, list) and source_lous:
        return ", ".join(f"#{lou}" for lou in source_lous if lou is not None)
    lou = entry.get("lou")
    if lou is None:
        return ""
    return f"#{lou}"


def bool_text(value: Any) -> str:
    return "是" if bool(value) else ""


def topic_defs_from_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    raw_topics = payload.get("topics")
    if isinstance(raw_topics, list):
        return [topic for topic in raw_topics if isinstance(topic, dict)]
    parsed_rule = payload.get("parsed_rule")
    if isinstance(parsed_rule, dict):
        raw_topics = parsed_rule.get("topics")
        if isinstance(raw_topics, list):
            return [topic for topic in raw_topics if isinstance(topic, dict)]
    return []


def topic_display_name(topic: dict[str, Any]) -> str:
    return str(topic.get("short_name") or topic.get("name") or topic.get("id") or "未命名主题")


def topic_sheet_label(topic: dict[str, Any]) -> str:
    return str(topic.get("short_name") or topic.get("name") or topic.get("id") or "主题")


def sanitize_sheet_title(value: str) -> str:
    cleaned = str(value).replace("\n", " ").replace("\r", " ").strip()
    for char in '[]:*?/\\':
        cleaned = cleaned.replace(char, "_")
    cleaned = " ".join(cleaned.split())
    return cleaned or "主题"


def unique_sheet_title(existing_titles: set[str], preferred: str) -> str:
    base = sanitize_sheet_title(preferred)
    candidate = base[:SHEET_TITLE_LIMIT]
    if candidate not in existing_titles:
        existing_titles.add(candidate)
        return candidate

    suffix_index = 2
    while True:
        suffix = f"_{suffix_index}"
        trimmed = base[: max(1, SHEET_TITLE_LIMIT - len(suffix))]
        candidate = f"{trimmed}{suffix}"
        if candidate not in existing_titles:
            existing_titles.add(candidate)
            return candidate
        suffix_index += 1


def ordered_entries(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def sort_key(entry: dict[str, Any]) -> tuple[int, int]:
        lou = entry.get("lou")
        entry_id = entry.get("id")
        return (
            lou if isinstance(lou, int) else 10**12,
            entry_id if isinstance(entry_id, int) else 10**12,
        )

    return sorted(entries, key=sort_key)


def field_headers_for_topic(topic_id: str, entries: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    seen = set()
    for header in TOPIC_FIELD_ORDER.get(topic_id, []):
        headers.append(header)
        seen.add(header)
    for entry in entries:
        fields = entry.get("fields")
        if not isinstance(fields, dict):
            continue
        for key in fields:
            key_text = str(key)
            if key_text not in seen:
                headers.append(key_text)
                seen.add(key_text)
    return headers


def field_value(entry: dict[str, Any], key: str) -> Any:
    fields = entry.get("fields")
    if not isinstance(fields, dict):
        return ""
    return fields.get(key, "")


def entry_url(entry: dict[str, Any]) -> str:
    url = entry.get("url")
    if isinstance(url, str) and url.strip():
        return url
    return post_url(entry.get("pid"))


def entry_notes(entry: dict[str, Any]) -> str:
    notes: list[str] = []
    classification_note = entry.get("classification_note")
    if isinstance(classification_note, str) and classification_note.strip():
        notes.append(classification_note.strip())
    proxy_reason = entry.get("proxy_reason")
    if entry.get("proxy_submission"):
        notes.append(f"代投{f'：{proxy_reason.strip()}' if isinstance(proxy_reason, str) and proxy_reason.strip() else ''}")
    duplicate_lous = entry.get("duplicate_lous")
    if entry.get("has_duplicate") and isinstance(duplicate_lous, list) and duplicate_lous:
        notes.append("重复楼层：" + ", ".join(f"#{lou}" for lou in duplicate_lous if lou is not None))
    confidence = entry.get("confidence")
    if isinstance(confidence, (int, float)) and confidence < 1:
        notes.append(f"置信度：{round(confidence * 100)}%")
    return join_lines(notes)


def weighted_target_info(entry: dict[str, Any]) -> tuple[str, Any, str]:
    weighted_target = entry.get("weighted_target")
    target_hint = entry.get("weight_target_hint")
    if not isinstance(weighted_target, dict):
        return (str(target_hint or ""), "", "")
    author = weighted_target.get("author") if isinstance(weighted_target.get("author"), dict) else {}
    return (
        str(weighted_target.get("content") or target_hint or ""),
        weighted_target.get("lou", ""),
        str(author.get("username") or ""),
    )


def build_topic_rows(topic: dict[str, Any], entries: list[dict[str, Any]]) -> list[list[Any]]:
    topic_id = str(topic.get("id") or "")
    sorted_entries = ordered_entries(entries)
    dynamic_field_headers = field_headers_for_topic(topic_id, sorted_entries)
    headers = [
        "序号",
        "投稿楼层",
        "发布时间",
        "作者",
        "UID",
        *dynamic_field_headers,
        "统计结果",
        "来源楼层",
        "加权目标",
        "加权目标楼层",
        "加权目标作者",
        "需要复核",
        "备注",
        "链接",
    ]
    rows: list[list[Any]] = [headers]
    for index, entry in enumerate(sorted_entries, start=1):
        author = entry.get("author") if isinstance(entry.get("author"), dict) else {}
        target_content, target_lou, target_author = weighted_target_info(entry)
        row = [
            index,
            entry.get("lou", ""),
            entry.get("postdate", ""),
            author.get("username", ""),
            author.get("uid", ""),
        ]
        row.extend(field_value(entry, field_name) for field_name in dynamic_field_headers)
        row.extend([
            entry.get("content", ""),
            source_lous_text(entry),
            target_content,
            target_lou,
            target_author,
            bool_text(entry.get("needs_manual_review")),
            entry_notes(entry),
            entry_url(entry),
        ])
        rows.append(row)
    return rows


def apply_sheet_formatting(worksheet: Worksheet, max_column_width: int) -> None:
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = max((len(value) for value in values[:200]), default=0) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), max_column_width)


def add_sheet(workbook: Workbook, title: str, rows: list[list[Any]], max_column_width: int) -> None:
    worksheet = workbook.create_sheet(title=title)
    for row in rows:
        worksheet.append([format_cell(value) for value in row])
    apply_sheet_formatting(worksheet, max_column_width)


def build_workbook(payload: dict[str, Any], max_column_width: int) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    raw_entries = payload.get("entries")
    entries = [entry for entry in raw_entries if isinstance(entry, dict)] if isinstance(raw_entries, list) else []
    entries_by_topic: dict[str, list[dict[str, Any]]] = {}
    for entry in entries:
        topic_id = str(entry.get("topic_id") or "unclassified")
        entries_by_topic.setdefault(topic_id, []).append(entry)

    topics = topic_defs_from_payload(payload)
    topic_ids_in_order: list[str] = []
    seen_topic_ids = set()
    for topic in topics:
        topic_id = str(topic.get("id") or "")
        if not topic_id or topic_id in seen_topic_ids:
            continue
        topic_ids_in_order.append(topic_id)
        seen_topic_ids.add(topic_id)

    for topic_id in sorted(entries_by_topic):
        if topic_id not in seen_topic_ids:
            topic_ids_in_order.append(topic_id)
            seen_topic_ids.add(topic_id)

    topics_by_id = {str(topic.get("id") or ""): topic for topic in topics if str(topic.get("id") or "")}
    existing_titles: set[str] = set()
    for index, topic_id in enumerate(topic_ids_in_order, start=1):
        topic = topics_by_id.get(topic_id) or {
            "id": topic_id,
            "name": topic_id,
            "short_name": topic_id,
        }
        sheet_name = unique_sheet_title(existing_titles, f"{index:02d}_{topic_sheet_label(topic)}")
        add_sheet(workbook, sheet_name, build_topic_rows(topic, entries_by_topic.get(topic_id, [])), max_column_width)

    if not workbook.sheetnames:
        fallback_title = unique_sheet_title(set(), "01_主题")
        add_sheet(workbook, fallback_title, [["序号", "统计结果"], ["", "没有可导出的安价结果"]], max_column_width)
    return workbook


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = (args.output.resolve() if args.output else default_output_path(input_path))

    payload = load_payload(input_path)
    workbook = build_workbook(payload, max_column_width=max(10, args.max_column_width))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    entries = payload.get("entries", []) if isinstance(payload.get("entries"), list) else []
    print(
        f"Exported {output_path} with {len(workbook.sheetnames)} topic sheets and {len(entries)} entries."
    )


if __name__ == "__main__":
    main()